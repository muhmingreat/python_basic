import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import os

wb = xl.load_workbook("C:/Users/ALFA/Desktop/transactions.xlsx")
print("Current working directory:", os.getcwd())


# wb = openpyxl.load_workbook("transactions.xlsx")
# sheet = wb["Sheet1"]

# for row in range(2, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     corrected_price = cell.value * 0.9
#     corrected_price_cell = sheet.cell(row, 4)
#     corrected_price_cell.value = corrected_price

# value = Reference(sheet,
#          min_row = 2,
#          max_row = sheet.max_row,
#          min_col = 4,
#          max_col = 4  )

# chart  = BarChart()
# chart.add_data(value)
# sheet.add_chart(chart, "e2")

# # cell = sheet["a1"]
# # cell = sheet.cell(row=1, column=1)
# # print(cell.value)

# wb.save('transactions3.xlsx')